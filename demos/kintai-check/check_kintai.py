"""
demos/kintai-check/
タイムカードExcel と 勤怠申請Excel を突合し、
承認可／要確認を仕分けしたHTMLレポートを生成する。

ロジック:
- 有給申請：対象日の打刻がなければ✅承認可、あれば⚠️要確認
- 残業申請：対象日の終業打刻から18:00を引いた超過時間と申請時間を比較（差15分以内で整合）
- 休出申請：申請時刻と実打刻を比較（開始・終了とも差15分以内で整合）
- 別枠：非営業日で申請されていない打刻（申請漏れ）を検出
"""
from openpyxl import load_workbook
from datetime import datetime, time, date, timedelta
from pathlib import Path

BASE = Path(__file__).parent
TC_PATH = BASE / "timecard.xlsx"
APP_PATH = BASE / "applications.xlsx"
OUT_PATH = BASE / "output" / "check-report.html"

DEFAULT_START = time(9, 0)
DEFAULT_END = time(18, 0)
TOLERANCE_MIN = 15  # 15分以内は整合と見なす


def load_timecard():
    wb = load_workbook(TC_PATH, data_only=True)
    ws = wb.active
    rows = []
    headers = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = row
            continue
        emp_id, name, d, wd, start_t, end_t, break_min, memo = row
        # 日付は文字列で格納（"YYYY-MM-DD"）
        if isinstance(d, datetime):
            d = d.date()
        elif isinstance(d, str):
            d = datetime.strptime(d, "%Y-%m-%d").date()
        rows.append({
            "emp_id": emp_id,
            "name": name,
            "date": d,
            "weekday": wd,
            "start": start_t,
            "end": end_t,
            "break_min": break_min or 0,
            "memo": memo or "",
        })
    return rows


def load_applications():
    wb = load_workbook(APP_PATH, data_only=True)
    ws = wb.active
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        app_id, emp_id, name, kind, d, start_t, end_t, hours, memo = row
        if isinstance(d, datetime):
            d = d.date()
        elif isinstance(d, str):
            d = datetime.strptime(d, "%Y-%m-%d").date()
        rows.append({
            "app_id": app_id,
            "emp_id": emp_id,
            "name": name,
            "kind": kind,
            "date": d,
            "start": start_t,
            "end": end_t,
            "hours": hours,
            "memo": memo or "",
        })
    return rows


def find_punch(timecard, emp_id, d):
    for r in timecard:
        if r["emp_id"] == emp_id and r["date"] == d:
            return r
    return None


def minutes_diff(t1, t2):
    """t1 - t2 を分単位で返す（両方timeオブジェクト）"""
    if t1 is None or t2 is None:
        return None
    dt1 = datetime.combine(date.today(), t1)
    dt2 = datetime.combine(date.today(), t2)
    return int((dt1 - dt2).total_seconds() / 60)


def format_time(t):
    if t is None:
        return "—"
    if isinstance(t, time):
        return t.strftime("%H:%M")
    return str(t)


def check_application(app, timecard):
    """申請1件を突合し、判定と根拠を返す"""
    punch = find_punch(timecard, app["emp_id"], app["date"])

    if app["kind"] == "有給休暇":
        if punch and punch["start"] is not None:
            return {
                "verdict": "warn",
                "title": "有給申請日に出勤打刻あり",
                "detail": f"有給申請（終日）に対し、{format_time(punch['start'])}-{format_time(punch['end'])}の打刻が残っています。有給を取り下げるか、実績を修正する必要があります。",
            }
        else:
            return {
                "verdict": "ok",
                "title": "打刻なし・整合",
                "detail": f"有給申請の対象日に打刻はありません。申請内容どおりです。",
            }

    if app["kind"] == "残業":
        if not punch or punch["end"] is None:
            return {
                "verdict": "warn",
                "title": "残業申請日に打刻なし",
                "detail": "残業申請に対し、その日の打刻が見当たりません。",
            }
        # 定時超過時間を計算
        over_min = minutes_diff(punch["end"], DEFAULT_END)
        applied_min = int(app["hours"] * 60) if app["hours"] else 0
        diff = over_min - applied_min
        if abs(diff) <= TOLERANCE_MIN:
            return {
                "verdict": "ok",
                "title": "申請時間と実績が整合",
                "detail": f"申請{applied_min // 60}時間{applied_min % 60:02d}分 vs 実打刻超過{over_min // 60}時間{over_min % 60:02d}分（差{abs(diff)}分、許容内）。",
            }
        else:
            return {
                "verdict": "warn",
                "title": "残業時間の差異",
                "detail": f"申請{applied_min // 60}時間{applied_min % 60:02d}分 vs 実打刻超過{over_min // 60}時間{over_min % 60:02d}分（差{abs(diff)}分）。申請修正または打刻の補正が必要です。",
            }

    if app["kind"] == "休日出勤":
        if not punch or punch["start"] is None:
            return {
                "verdict": "warn",
                "title": "休出申請日に打刻なし",
                "detail": "休日出勤申請に対し、その日の打刻が見当たりません。",
            }
        start_diff = minutes_diff(punch["start"], app["start"])
        end_diff = minutes_diff(punch["end"], app["end"])
        if abs(start_diff) <= TOLERANCE_MIN and abs(end_diff) <= TOLERANCE_MIN:
            return {
                "verdict": "ok",
                "title": "申請時刻と実打刻が整合",
                "detail": f"申請{format_time(app['start'])}-{format_time(app['end'])} vs 実打刻{format_time(punch['start'])}-{format_time(punch['end'])}（差±{max(abs(start_diff), abs(end_diff))}分、許容内）。",
            }
        else:
            return {
                "verdict": "warn",
                "title": "休出時刻の差異",
                "detail": f"申請{format_time(app['start'])}-{format_time(app['end'])} vs 実打刻{format_time(punch['start'])}-{format_time(punch['end'])}（開始差{start_diff:+d}分・終了差{end_diff:+d}分）。",
            }

    return {"verdict": "warn", "title": "未対応の申請種別", "detail": app["kind"]}


def detect_missing_applications(timecard, applications):
    """申請されていない非営業日打刻を検出"""
    applied_dates = {(a["emp_id"], a["date"]) for a in applications}
    missing = []
    for r in timecard:
        if r["start"] is None:
            continue
        # 非営業日判定：曜日が土日
        if r["weekday"] in ("土", "日"):
            key = (r["emp_id"], r["date"])
            if key not in applied_dates:
                missing.append(r)
    return missing


def render_html(applications, results, missing):
    ok_count = sum(1 for r in results if r["verdict"] == "ok")
    warn_count = sum(1 for r in results if r["verdict"] == "warn")

    # 申請明細行
    rows_html = []
    for app, res in zip(applications, results):
        badge_class = "badge-ok" if res["verdict"] == "ok" else "badge-warn"
        badge_icon = "✅" if res["verdict"] == "ok" else "⚠️"
        badge_text = "承認可" if res["verdict"] == "ok" else "要確認"
        applied_range = ""
        if app["kind"] == "有給休暇":
            applied_range = "終日"
        elif app["start"] and app["end"]:
            applied_range = f"{format_time(app['start'])}〜{format_time(app['end'])}"
        rows_html.append(f"""
          <tr class="row-{res['verdict']}">
            <td class="cell-id">{app['app_id']}</td>
            <td>{app['name']}<br><span class="muted">{app['emp_id']}</span></td>
            <td>{app['kind']}</td>
            <td>{app['date'].strftime('%Y-%m-%d')}<br><span class="muted">{applied_range}</span></td>
            <td><span class="badge {badge_class}">{badge_icon} {badge_text}</span></td>
            <td class="cell-detail"><strong>{res['title']}</strong><br>{res['detail']}</td>
          </tr>""")

    # 申請漏れ行
    missing_html = ""
    if missing:
        miss_rows = []
        for m in missing:
            miss_rows.append(f"""
              <tr class="row-warn">
                <td>{m['name']}<br><span class="muted">{m['emp_id']}</span></td>
                <td>{m['date'].strftime('%Y-%m-%d')}（{m['weekday']}）</td>
                <td>{format_time(m['start'])}〜{format_time(m['end'])}</td>
                <td><strong>休日出勤申請なしで打刻あり</strong><br>非営業日の打刻に対して申請が見当たりません。本人と上長に事後申請の要否を確認してください。</td>
              </tr>""")
        missing_html = f"""
    <section class="section">
      <h2>申請漏れの疑い（打刻側から検出）</h2>
      <p class="hint">非営業日（土日・祝日）に打刻があり、対応する休日出勤申請が見当たらない行です。</p>
      <table class="detail-table">
        <thead>
          <tr><th>氏名</th><th>対象日</th><th>実打刻</th><th>検出内容</th></tr>
        </thead>
        <tbody>{''.join(miss_rows)}</tbody>
      </table>
    </section>"""

    html = f"""<!DOCTYPE html>
<html lang="ja">
<head>
<meta charset="UTF-8">
<title>勤怠突合チェックレポート｜2026年3月度｜○△商事</title>
<style>
  * {{ box-sizing: border-box; }}
  body {{ font-family: "Hiragino Kaku Gothic ProN", "Noto Sans JP", sans-serif;
         margin: 0; padding: 0; background: #f5f6fa; color: #1f2330; }}
  .container {{ max-width: 1100px; margin: 0 auto; padding: 40px 24px; }}
  .hero {{ background: linear-gradient(135deg, #2d4e8c, #1a2f57); color: #fff;
          padding: 36px 40px; border-radius: 14px; box-shadow: 0 6px 20px rgba(0,0,0,.1); }}
  .hero h1 {{ margin: 0 0 6px; font-size: 26px; letter-spacing: .02em; }}
  .hero .sub {{ opacity: .85; font-size: 14px; }}
  .kpi-bar {{ display: grid; grid-template-columns: repeat(4, 1fr); gap: 14px;
             margin-top: -22px; padding: 0 16px; position: relative; z-index: 2; }}
  .kpi {{ background: #fff; border-radius: 12px; padding: 18px 20px;
         box-shadow: 0 4px 14px rgba(0,0,0,.06); }}
  .kpi .label {{ font-size: 12px; color: #6b7280; margin-bottom: 6px; letter-spacing: .04em; }}
  .kpi .value {{ font-size: 30px; font-weight: 700; line-height: 1; }}
  .kpi.ok .value {{ color: #15803d; }}
  .kpi.warn .value {{ color: #c2410c; }}
  .kpi.miss .value {{ color: #7c3aed; }}
  .kpi.total .value {{ color: #1f2937; }}
  .section {{ background: #fff; border-radius: 12px; padding: 28px 32px; margin-top: 28px;
             box-shadow: 0 2px 10px rgba(0,0,0,.04); }}
  .section h2 {{ margin: 0 0 4px; font-size: 18px; color: #1f2937;
                border-left: 4px solid #2d4e8c; padding-left: 12px; }}
  .hint {{ font-size: 13px; color: #6b7280; margin: 6px 0 18px; }}
  table {{ width: 100%; border-collapse: collapse; font-size: 14px; }}
  th {{ background: #f3f4f6; color: #4b5563; text-align: left; padding: 10px 12px;
        border-bottom: 2px solid #e5e7eb; font-weight: 600; }}
  td {{ padding: 12px; border-bottom: 1px solid #e5e7eb; vertical-align: top; }}
  .row-warn {{ background: #fffbeb; }}
  .row-warn td.cell-detail {{ color: #9a3412; }}
  .cell-id {{ font-family: "SF Mono", Menlo, monospace; font-size: 12px; color: #6b7280; }}
  .muted {{ color: #9ca3af; font-size: 11px; }}
  .badge {{ display: inline-block; padding: 4px 10px; border-radius: 999px;
            font-size: 12px; font-weight: 600; letter-spacing: .02em; }}
  .badge-ok {{ background: #dcfce7; color: #15803d; }}
  .badge-warn {{ background: #ffedd5; color: #c2410c; }}
  .footer {{ text-align: center; color: #9ca3af; font-size: 12px; padding: 26px 0; }}
  .summary-note {{ background: #eff6ff; border-left: 4px solid #3b82f6;
                   padding: 14px 18px; margin-top: 16px; border-radius: 6px;
                   font-size: 13px; color: #1e3a8a; }}
</style>
</head>
<body>
  <div class="container">
    <div class="hero">
      <h1>勤怠突合チェックレポート</h1>
      <div class="sub">対象期間：2026年3月｜会社：○△商事（架空）｜対象申請：10件</div>
    </div>

    <div class="kpi-bar">
      <div class="kpi total"><div class="label">申請総数</div><div class="value">{len(applications)}</div></div>
      <div class="kpi ok"><div class="label">✅ 承認可</div><div class="value">{ok_count}</div></div>
      <div class="kpi warn"><div class="label">⚠️ 要確認</div><div class="value">{warn_count}</div></div>
      <div class="kpi miss"><div class="label">申請漏れ疑い</div><div class="value">{len(missing)}</div></div>
    </div>

    <section class="section">
      <h2>申請明細（承認可・要確認の仕分け）</h2>
      <p class="hint">申請10件それぞれに対し、タイムカード上の打刻と突合した結果を示しています。要確認は人間側で最終判定をお願いします。</p>
      <table class="detail-table">
        <thead>
          <tr>
            <th style="width:70px;">申請ID</th>
            <th style="width:130px;">氏名</th>
            <th style="width:90px;">種別</th>
            <th style="width:140px;">対象日／範囲</th>
            <th style="width:90px;">判定</th>
            <th>差分根拠</th>
          </tr>
        </thead>
        <tbody>{''.join(rows_html)}</tbody>
      </table>
    </section>{missing_html}

    <section class="section">
      <h2>運用メモ</h2>
      <div class="summary-note">
        <strong>100%自動化は幻想。80%自動仕分け＋20%人間判定が現実解。</strong><br>
        AIは時間差・日付差・打刻有無のフラグ立てを引き受け、最終承認印は人間が押す運用を推奨します。
        月100件×5分＝500分の目視突合を、要確認数件×6分＝約30分に圧縮できる想定です。
      </div>
    </section>

    <div class="footer">
      Generated by Claude Code｜データはすべて架空、機微情報（マイナンバー・住所・生年月日）は題材から除外済み
    </div>
  </div>
</body>
</html>
"""
    return html


def main():
    timecard = load_timecard()
    applications = load_applications()

    results = [check_application(a, timecard) for a in applications]
    missing = detect_missing_applications(timecard, applications)

    ok_count = sum(1 for r in results if r["verdict"] == "ok")
    warn_count = sum(1 for r in results if r["verdict"] == "warn")

    print(f"申請総数: {len(applications)}")
    print(f"✅ 承認可: {ok_count}")
    print(f"⚠️ 要確認: {warn_count}")
    print(f"申請漏れ疑い: {len(missing)}")
    print()
    for app, res in zip(applications, results):
        mark = "✅" if res["verdict"] == "ok" else "⚠️"
        print(f"  {mark} [{app['app_id']}] {app['name']} / {app['kind']} / {app['date']} → {res['title']}")
    for m in missing:
        print(f"  🔶 [申請漏れ] {m['name']} / {m['date']}（{m['weekday']}） / 打刻 {format_time(m['start'])}-{format_time(m['end'])}")

    OUT_PATH.parent.mkdir(exist_ok=True)
    html = render_html(applications, results, missing)
    OUT_PATH.write_text(html, encoding="utf-8")
    print(f"\nHTMLレポート: {OUT_PATH}")


if __name__ == "__main__":
    main()
